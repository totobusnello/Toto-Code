#!/usr/bin/env python3
"""
Extract financial data from Excel files (financial models, reports, data exports).
Supports multiple sheets, named ranges, and pattern-based metric extraction.
"""

import sys
import json
import re
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl --break-system-packages", file=sys.stderr)
    sys.exit(1)


def extract_all_sheets(excel_path, include_formulas=False):
    """Extract data from all sheets in the workbook."""
    workbook = openpyxl.load_workbook(excel_path, data_only=not include_formulas)
    
    sheets_data = {}
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        # Extract data
        data = []
        for row in sheet.iter_rows(values_only=True):
            # Convert row to list, handling None values
            row_data = [cell if cell is not None else "" for cell in row]
            data.append(row_data)
        
        # Get max dimensions
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        sheets_data[sheet_name] = {
            "dimensions": {
                "rows": max_row,
                "columns": max_col
            },
            "data": data
        }
    
    return sheets_data


def extract_financial_metrics(excel_path, metric_keywords=None):
    """Extract financial metrics by searching for keywords in the first column/row."""
    if metric_keywords is None:
        metric_keywords = [
            "revenue", "sales", "ebitda", "ebit", "net income", "gross profit",
            "operating income", "cash flow", "capex", "net debt", "equity value",
            "enterprise value", "arr", "mrr", "gross margin", "operating margin",
            "employees", "customers", "churn", "cac", "ltv"
        ]
    
    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    
    extracted_metrics = {}
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet_metrics = {}
        
        # Search through rows
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row), start=1):
            first_cell = row[0].value
            
            if first_cell and isinstance(first_cell, str):
                first_cell_lower = first_cell.lower().strip()
                
                # Check if any keyword matches
                for keyword in metric_keywords:
                    if keyword.lower() in first_cell_lower:
                        # Extract values from this row
                        values = []
                        for cell in row[1:]:
                            if cell.value is not None:
                                values.append(cell.value)
                        
                        if values:
                            if keyword not in sheet_metrics:
                                sheet_metrics[keyword] = []
                            
                            sheet_metrics[keyword].append({
                                "label": first_cell,
                                "row": row_idx,
                                "values": values
                            })
        
        if sheet_metrics:
            extracted_metrics[sheet_name] = sheet_metrics
    
    return extracted_metrics


def extract_tables(excel_path, sheet_name=None):
    """Extract tables from Excel (contiguous data regions)."""
    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    
    sheets_to_process = [sheet_name] if sheet_name else workbook.sheetnames
    tables_data = {}
    
    for sname in sheets_to_process:
        sheet = workbook[sname]
        
        # Find tables by looking for contiguous data regions
        tables = []
        
        # Simple table detection: look for rows with headers followed by data
        for start_row in range(1, min(sheet.max_row, 50)):  # Check first 50 rows
            first_row = list(sheet.iter_rows(min_row=start_row, max_row=start_row, values_only=True))[0]
            
            # Check if this looks like a header row (mostly text)
            if first_row and any(isinstance(cell, str) and cell.strip() for cell in first_row):
                # Count consecutive rows with data
                data_rows = []
                for row_idx in range(start_row, min(sheet.max_row + 1, start_row + 1000)):
                    row_data = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
                    
                    if any(cell is not None and cell != "" for cell in row_data):
                        data_rows.append(row_data)
                    elif len(data_rows) > 2:  # At least header + 2 data rows
                        break
                
                if len(data_rows) > 2:
                    tables.append({
                        "start_row": start_row,
                        "end_row": start_row + len(data_rows) - 1,
                        "headers": data_rows[0],
                        "data": data_rows[1:]
                    })
        
        if tables:
            tables_data[sname] = tables
    
    return tables_data


def search_values(excel_path, search_pattern):
    """Search for values matching a regex pattern across all sheets."""
    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    
    results = {}
    pattern = re.compile(search_pattern, re.IGNORECASE)
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        matches = []
        
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if pattern.search(cell.value):
                        matches.append({
                            "cell": cell.coordinate,
                            "value": cell.value,
                            "row": cell.row,
                            "column": get_column_letter(cell.column)
                        })
        
        if matches:
            results[sheet_name] = matches
    
    return results


def main():
    if len(sys.argv) < 2:
        print("Usage: python extract_excel_financials.py <excel_path> [--mode all|metrics|tables|sheets|search] [--search 'pattern'] [--sheet 'name']", file=sys.stderr)
        sys.exit(1)
    
    excel_path = sys.argv[1]
    mode = "all"
    search_pattern = None
    sheet_name = None
    
    # Parse arguments
    if "--mode" in sys.argv:
        mode_idx = sys.argv.index("--mode")
        if mode_idx + 1 < len(sys.argv):
            mode = sys.argv[mode_idx + 1]
    
    if "--search" in sys.argv:
        search_idx = sys.argv.index("--search")
        if search_idx + 1 < len(sys.argv):
            search_pattern = sys.argv[search_idx + 1]
            mode = "search"
    
    if "--sheet" in sys.argv:
        sheet_idx = sys.argv.index("--sheet")
        if sheet_idx + 1 < len(sys.argv):
            sheet_name = sys.argv[sheet_idx + 1]
    
    if not Path(excel_path).exists():
        print(f"Error: File not found: {excel_path}", file=sys.stderr)
        sys.exit(1)
    
    results = {}
    
    try:
        if mode in ["sheets", "all"]:
            results["sheets"] = extract_all_sheets(excel_path)
        
        if mode in ["metrics", "all"]:
            results["financial_metrics"] = extract_financial_metrics(excel_path)
        
        if mode in ["tables", "all"]:
            results["tables"] = extract_tables(excel_path, sheet_name)
        
        if mode == "search" and search_pattern:
            results["search_results"] = search_values(excel_path, search_pattern)
        
        # Output results
        print(json.dumps(results, indent=2, default=str))
        
    except Exception as e:
        print(f"Error processing Excel file: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
