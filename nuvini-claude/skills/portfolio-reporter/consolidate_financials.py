"""
Financial Consolidation Module for Nuvini Portfolio

Consolidates financial data from multiple companies into a single summary,
calculating portfolio-level metrics, margins, and growth rates.
"""

import pandas as pd
import numpy as np
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass, field
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from extract_financials import (
    CompanyData, MonthlyFinancials, KPIData,
    format_currency, format_percentage, format_growth
)


@dataclass
class PortfolioSummary:
    """Consolidated portfolio financial summary"""
    companies: List[str]
    report_period: str  # e.g., "Jan-Jul 2025" or "Q3 2025"
    year: int
    end_month: int

    # Aggregated financials by period
    consolidated_monthly: Dict[str, MonthlyFinancials] = field(default_factory=dict)

    # Company-level summaries
    company_ytd: Dict[str, MonthlyFinancials] = field(default_factory=dict)
    company_kpis: Dict[str, KPIData] = field(default_factory=dict)

    # Portfolio KPIs
    portfolio_rule_of_40: float = 0.0
    avg_ebitda_margin: float = 0.0
    avg_yoy_growth: float = 0.0
    companies_above_40: int = 0
    positive_fcf_companies: int = 0

    # MoM changes
    current_month_revenue: float = 0.0
    previous_month_revenue: float = 0.0
    mom_revenue_change: float = 0.0
    current_month_ebitda: float = 0.0
    previous_month_ebitda: float = 0.0
    mom_ebitda_change: float = 0.0
    companies_with_positive_growth: int = 0

    # Quarterly data (Phase A: US-003)
    is_quarterly: bool = False
    quarter: Optional[int] = None
    start_month: int = 1

    # Quarterly aggregated financials
    quarterly_financials: Dict[str, MonthlyFinancials] = field(default_factory=dict)  # Q1, Q2, Q3, Q4

    # QoQ changes
    current_quarter_revenue: float = 0.0
    previous_quarter_revenue: float = 0.0
    qoq_revenue_change: Optional[float] = None
    current_quarter_ebitda: float = 0.0
    previous_quarter_ebitda: float = 0.0
    qoq_ebitda_change: Optional[float] = None

    # Company quarterly data
    company_quarterly: Dict[str, MonthlyFinancials] = field(default_factory=dict)  # company_name -> quarter totals


def consolidate_portfolio(
    company_data: Dict[str, CompanyData],
    year: int,
    end_month: int
) -> PortfolioSummary:
    """
    Consolidate financial data from multiple companies into a portfolio summary.

    Args:
        company_data: Dictionary mapping company names to their CompanyData
        year: Report year (e.g., 2025)
        end_month: Last month to include (e.g., 7 for July)

    Returns:
        PortfolioSummary with all consolidated metrics
    """
    month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                   'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

    summary = PortfolioSummary(
        companies=list(company_data.keys()),
        report_period=f"Jan-{month_names[end_month-1]} {year}",
        year=year,
        end_month=end_month
    )

    # Calculate YTD for each company
    for company_name, data in company_data.items():
        ytd = data.get_ytd_totals(year, end_month)
        summary.company_ytd[company_name] = ytd

        # Get latest KPIs
        kpi = data.get_latest_kpis(year, end_month)
        if kpi:
            summary.company_kpis[company_name] = kpi

    # Consolidate monthly data
    for month in range(1, end_month + 1):
        period = f"{year}{month:02d}"
        consolidated = MonthlyFinancials(period=period)

        for company_name, data in company_data.items():
            if period in data.monthly_financials:
                monthly = data.monthly_financials[period]
                consolidated.gross_revenue += monthly.gross_revenue
                consolidated.net_revenue += monthly.net_revenue
                consolidated.deductions += monthly.deductions
                consolidated.gross_profit += monthly.gross_profit
                consolidated.operating_costs += monthly.operating_costs
                consolidated.operating_expenses += monthly.operating_expenses
                consolidated.ebitda += monthly.ebitda
                consolidated.depreciation_amortization += monthly.depreciation_amortization
                consolidated.net_profit += monthly.net_profit
                consolidated.free_cash_flow += monthly.free_cash_flow

        # Calculate margins
        if consolidated.net_revenue > 0:
            consolidated.gross_margin = consolidated.gross_profit / consolidated.net_revenue
            consolidated.ebitda_margin = consolidated.ebitda / consolidated.net_revenue
            consolidated.net_margin = consolidated.net_profit / consolidated.net_revenue

        summary.consolidated_monthly[period] = consolidated

    # Calculate portfolio YTD totals
    portfolio_ytd = MonthlyFinancials(period=f"{year}01-{year}{end_month:02d}")
    for period, monthly in summary.consolidated_monthly.items():
        portfolio_ytd.gross_revenue += monthly.gross_revenue
        portfolio_ytd.net_revenue += monthly.net_revenue
        portfolio_ytd.gross_profit += monthly.gross_profit
        portfolio_ytd.ebitda += monthly.ebitda
        portfolio_ytd.net_profit += monthly.net_profit
        portfolio_ytd.free_cash_flow += monthly.free_cash_flow

    if portfolio_ytd.net_revenue > 0:
        portfolio_ytd.gross_margin = portfolio_ytd.gross_profit / portfolio_ytd.net_revenue
        portfolio_ytd.ebitda_margin = portfolio_ytd.ebitda / portfolio_ytd.net_revenue

    summary.company_ytd['PORTFOLIO'] = portfolio_ytd

    # Calculate MoM changes
    current_period = f"{year}{end_month:02d}"
    previous_period = f"{year}{end_month-1:02d}" if end_month > 1 else f"{year-1}12"

    if current_period in summary.consolidated_monthly:
        summary.current_month_revenue = summary.consolidated_monthly[current_period].net_revenue
        summary.current_month_ebitda = summary.consolidated_monthly[current_period].ebitda

    if previous_period in summary.consolidated_monthly:
        summary.previous_month_revenue = summary.consolidated_monthly[previous_period].net_revenue
        summary.previous_month_ebitda = summary.consolidated_monthly[previous_period].ebitda

    if summary.previous_month_revenue > 0:
        summary.mom_revenue_change = (
            summary.current_month_revenue - summary.previous_month_revenue
        ) / summary.previous_month_revenue

    if summary.previous_month_ebitda > 0:
        summary.mom_ebitda_change = (
            summary.current_month_ebitda - summary.previous_month_ebitda
        ) / summary.previous_month_ebitda

    # Calculate company-level MoM and count positive growth
    for company_name, data in company_data.items():
        current_rev = 0.0
        previous_rev = 0.0

        if current_period in data.monthly_financials:
            current_rev = data.monthly_financials[current_period].net_revenue
        if previous_period in data.monthly_financials:
            previous_rev = data.monthly_financials[previous_period].net_revenue

        if previous_rev > 0 and current_rev > previous_rev:
            summary.companies_with_positive_growth += 1

    # Calculate portfolio KPIs
    ebitda_margins = []
    yoy_growths = []
    rule_of_40_scores = []

    for company_name, kpi in summary.company_kpis.items():
        if kpi.ebitda_margin > 0:
            ebitda_margins.append(kpi.ebitda_margin)
        if kpi.yoy_growth != 0:
            yoy_growths.append(kpi.yoy_growth)
        if kpi.rule_of_40 > 0:
            rule_of_40_scores.append(kpi.rule_of_40)
            if kpi.rule_of_40 >= 40:
                summary.companies_above_40 += 1

    # Also check FCF for positive FCF count
    for company_name, ytd in summary.company_ytd.items():
        if company_name != 'PORTFOLIO' and ytd.free_cash_flow > 0:
            summary.positive_fcf_companies += 1

    if ebitda_margins:
        summary.avg_ebitda_margin = sum(ebitda_margins) / len(ebitda_margins)
    if yoy_growths:
        summary.avg_yoy_growth = sum(yoy_growths) / len(yoy_growths)
    if rule_of_40_scores:
        summary.portfolio_rule_of_40 = sum(rule_of_40_scores) / len(rule_of_40_scores)

    return summary


def create_consolidated_excel(
    summary: PortfolioSummary,
    output_path: str
) -> str:
    """
    Create the consolidated Excel summary file.

    Args:
        summary: PortfolioSummary with all consolidated data
        output_path: Path to save the Excel file

    Returns:
        Path to the created file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Combinação"

    # Styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E3A8A")
    section_font = Font(bold=True, size=11)
    number_font = Font(size=10)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws['A1'] = "Consolidated Financial Summary"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:H1')

    ws['A2'] = f"Report Period: {summary.report_period}"
    ws['A2'].font = Font(size=12)

    # Create period columns
    periods = sorted(summary.consolidated_monthly.keys())
    col_offset = 4  # Start data from column E

    # Header row
    row = 4
    ws.cell(row=row, column=1, value="Category")
    ws.cell(row=row, column=2, value="Account")
    ws.cell(row=row, column=3, value="Unit")

    for idx, period in enumerate(periods):
        ws.cell(row=row, column=col_offset + idx, value=period)
        ws.cell(row=row, column=col_offset + idx).font = header_font
        ws.cell(row=row, column=col_offset + idx).fill = header_fill

    # DRE Section
    row = 6
    ws.cell(row=row, column=2, value="DRE (Income Statement)")
    ws.cell(row=row, column=2).font = section_font

    # Income Statement line items
    line_items = [
        ("", "Gross Revenue", "R$", "gross_revenue"),
        ("", "Deductions", "R$", "deductions"),
        ("", "Net Revenue", "R$", "net_revenue"),
        ("", "Gross Profit", "R$", "gross_profit"),
        ("", "EBITDA", "R$", "ebitda"),
        ("", "Net Profit", "R$", "net_profit"),
        ("", "Gross Margin", "%", "gross_margin"),
        ("", "EBITDA Margin", "%", "ebitda_margin"),
        ("", "Net Margin", "%", "net_margin"),
    ]

    for item in line_items:
        row += 1
        ws.cell(row=row, column=1, value=item[0])
        ws.cell(row=row, column=2, value=item[1])
        ws.cell(row=row, column=3, value=item[2])

        for idx, period in enumerate(periods):
            monthly = summary.consolidated_monthly.get(period)
            if monthly:
                value = getattr(monthly, item[3], 0)
                if item[2] == "%":
                    ws.cell(row=row, column=col_offset + idx, value=value)
                    ws.cell(row=row, column=col_offset + idx).number_format = '0.0%'
                else:
                    ws.cell(row=row, column=col_offset + idx, value=value / 1000)  # In thousands
                    ws.cell(row=row, column=col_offset + idx).number_format = '#,##0'

    # Company Summary Section
    row += 3
    ws.cell(row=row, column=2, value="Company YTD Summary")
    ws.cell(row=row, column=2).font = section_font

    row += 1
    headers = ["Company", "Net Revenue", "EBITDA", "EBITDA Margin", "Gross Margin", "FCF", "Rule of 40", "Status"]
    for idx, header in enumerate(headers):
        ws.cell(row=row, column=idx + 1, value=header)
        ws.cell(row=row, column=idx + 1).font = header_font
        ws.cell(row=row, column=idx + 1).fill = header_fill

    for company_name in summary.companies:
        row += 1
        ytd = summary.company_ytd.get(company_name)
        kpi = summary.company_kpis.get(company_name)

        ws.cell(row=row, column=1, value=company_name)

        if ytd:
            ws.cell(row=row, column=2, value=ytd.net_revenue / 1_000_000)
            ws.cell(row=row, column=2).number_format = '#,##0.0"M"'

            ws.cell(row=row, column=3, value=ytd.ebitda / 1_000_000)
            ws.cell(row=row, column=3).number_format = '#,##0.0"M"'

            ws.cell(row=row, column=4, value=ytd.ebitda_margin)
            ws.cell(row=row, column=4).number_format = '0.0%'

            ws.cell(row=row, column=5, value=ytd.gross_margin)
            ws.cell(row=row, column=5).number_format = '0.0%'

            ws.cell(row=row, column=6, value=ytd.free_cash_flow / 1_000_000)
            ws.cell(row=row, column=6).number_format = '#,##0.0"M"'

        if kpi:
            ws.cell(row=row, column=7, value=kpi.rule_of_40)
            ws.cell(row=row, column=7).number_format = '0.0'

            # Status based on Rule of 40 and growth
            if kpi.rule_of_40 >= 60:
                status = "Strong"
            elif kpi.rule_of_40 >= 40:
                status = "Growth"
            elif kpi.yoy_growth < 0:
                status = "Turnaround" if ytd and ytd.ebitda_margin > 0.35 else "Recovery"
            else:
                status = "Developing"
            ws.cell(row=row, column=8, value=status)

    # Portfolio totals row
    row += 1
    portfolio_ytd = summary.company_ytd.get('PORTFOLIO')
    ws.cell(row=row, column=1, value="TOTAL PORTFOLIO")
    ws.cell(row=row, column=1).font = Font(bold=True)

    if portfolio_ytd:
        ws.cell(row=row, column=2, value=portfolio_ytd.net_revenue / 1_000_000)
        ws.cell(row=row, column=2).number_format = '#,##0.0"M"'
        ws.cell(row=row, column=2).font = Font(bold=True)

        ws.cell(row=row, column=3, value=portfolio_ytd.ebitda / 1_000_000)
        ws.cell(row=row, column=3).number_format = '#,##0.0"M"'
        ws.cell(row=row, column=3).font = Font(bold=True)

        ws.cell(row=row, column=4, value=portfolio_ytd.ebitda_margin)
        ws.cell(row=row, column=4).number_format = '0.0%'
        ws.cell(row=row, column=4).font = Font(bold=True)

        ws.cell(row=row, column=5, value=portfolio_ytd.gross_margin)
        ws.cell(row=row, column=5).number_format = '0.0%'
        ws.cell(row=row, column=5).font = Font(bold=True)

        ws.cell(row=row, column=6, value=portfolio_ytd.free_cash_flow / 1_000_000)
        ws.cell(row=row, column=6).number_format = '#,##0.0"M"'
        ws.cell(row=row, column=6).font = Font(bold=True)

    ws.cell(row=row, column=7, value=summary.portfolio_rule_of_40)
    ws.cell(row=row, column=7).number_format = '0.0'
    ws.cell(row=row, column=7).font = Font(bold=True)

    ws.cell(row=row, column=8, value="Strong" if summary.portfolio_rule_of_40 >= 40 else "Developing")

    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 8
    for col in range(col_offset, col_offset + len(periods)):
        ws.column_dimensions[get_column_letter(col)].width = 12

    # Save
    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    # Test consolidation with sample data
    from extract_financials import extract_company_data

    # This would be called with actual file paths
    print("Consolidation module ready for use")
